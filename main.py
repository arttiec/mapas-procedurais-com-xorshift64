import tkinter as tk
import numpy as np
import pandas as pd
from time import time
from openpyxl import load_workbook
from xorshift import Xorshift64
from config import create_archive


def get_data():
    
    '''
    Carrega todos os dados do arquivo .xlsx e os retorna em formato de lista.

    Usamos o método load_workbook da biblioteca openpyxl para carrega a planilha para a variável wb
    e depois usamos a planilha ativa dentro do arquivo (wb.active).
    Então iteramos sobre cada linha do arquivo e adicionamos os dados à lista data,
    usamos o parâmetro values_only para informar ao iter_rows() que queremos apenas o conteúdo das células.

    Args:
        None

    Returns:
        data: lista com todos os dados da planilha seed.xlsx;
    '''

    wb = load_workbook('./seed.xlsx')
    spreadsheet = wb.active

    data = []

    for row in spreadsheet.iter_rows(values_only=True):
        data.append(row)

    return data


def saved_seeds_screen():
    
    '''
    Cria uma tela para exibir os dados obtidos a partir da função 'get_data()'.

    Criamos uma listbox para exibir os dados da planilha. Primeiramente, limpamos
    toda a listbox (listbox.delete(0, tk.END)) para garantir que a lista esteja vazia
    antes de adicionar novos dados. Logo após, usamos um loop para iterar sobre os dados
    obtidos e os adicionamos na listbox.

    Args:
        None

    Returns:
        None   
    '''
    
    width = 600
    height = 500

    window = tk.Tk();
    window.geometry(f'{width}x{height}')

    listbox = tk.Listbox(window)

    data = get_data()
    listbox.delete(0, tk.END)

    for row in data:
        listbox.insert(tk.END, row)
    
    listbox.pack(pady=10, fill=tk.BOTH, expand=True)
    
    window.mainloop()


def data_exist():

    '''
    Verifica se os dados que o usuário quer salvar já não existe dentro da planilha seed.xlsx.

    Verificamos isso através de uma condição: Caso a seed esteja na lista de valores pertencentes a coluna Seed
    e os steps estejam na lista de valores pertencentes a coluna Step, o mapa que o cliente deseja salvar
    já consta no arquivo.

    Args:
        None

    Returns:
        True (Boolean): caso dado existe dentro da planilha função retorna um valor verdadeiro 
    '''

    df = pd.read_excel('./seed.xlsx')

    if (seed in df['Seed'].tolist()) and (steps in df['Step'].tolist()):
        return True
    
    
def save_seed():

    '''
    Guarda a semente que o usuário deseja salvar no arquivo seed.xlsx.

    Primeiro, temos um try-except para tratar de uma excessão de NameError(variável não declarada)
    que ocorre quando o usuário está tentando salvar um valor que ainda não foi passado para os parâmetros seed e step.
    Caso a exceção não ocorra a função verifica se o dados existe dentro da planilha: Caso sim,
    ele exibi uma mensagem informando que valor ja existe, caso não, a função carrega a base de dados 
    e cria um dicionário para armazenar os dados que desejamos inserir, depois os 
    converter num DataFrame e concatena com nossa base de dados, adicionando-os ao arquivo existente.

    Args:
        None

    Returns:
        None 
    '''

    try:
        if not data_exist():
            df = pd.read_excel('./seed.xlsx')

            new_data = {'Name': name, 'Seed': seed, 'Step': steps}
            new_df = pd.DataFrame([new_data])

            df = pd.concat([df, new_df], ignore_index=True)

            df.to_excel('./seed.xlsx', index=False)
        
        else:
            print('Seed is already registered in seed.xlsx')

    except NameError:
        print('Seed or step variable is empty')



def save_screen():
    
    '''
    Exibi uma tela com um campo 'name' para que o usuário escolha, opcionalmente, um nome para a seed que deseja salvar.

    O botão contêm uma função lambda responsável por chamar, sequencialmente, as funções get_name_entry(),
    save_seed() e window.destroy(). Chamamos essa última para que a janela seja imediatamente fechada após
    ocorrer o salvamento dos dados.

    Args:
        None

    Returns:
        None 
    '''
    
    def get_name_entry():
        
        '''
        Pega o valor do campo 'name'.

        Args:
            None

        Returns:
            None
        '''                                                 
        
        global name

        name = save_entry.get()
        name = str(name)


    width = 400
    height = 150

    window = tk.Tk()
    window.geometry(f'{width}x{height}')
    window.title('Save Seed')

    label1 = tk.Label(window, text='Choose a name')
    label1.pack()

    save_entry = tk.Entry(window)
    save_entry.pack(pady=10)

    save_button = tk.Button(window, text='Save Seed', command=lambda: (get_name_entry(), save_seed(), window.destroy()))
    save_button.pack(pady=10)

    window.mainloop()

    
def choose_direction(generator):

    '''
    Escolha a direção para a qual o caminho do mapa será criado.

    Para tal escolha, sorteamos um número aleatório entre 0 e 3 correspondente a direção a se movimentar,
    para direita, cima, esquerda e baixo, respectivamente.

    Args:
        generator (Xorshift64): o gerador de números aleatórios.

    Returns:
        x (int): movimentação no eixo x;
        y (int): movimentação no eixo y.
    '''

    direction = generator.xorshift64() % 4

    if (direction == 0):
        x, y = 1, 0 

    elif (direction == 1):
        x, y = 0, -1 

    elif (direction == 2):
        x, y = -1, 0 

    elif (direction == 3):
        x, y = 0, 1 

    return x, y


def generate_room(rows, columns):
    
    '''
    Gera uma matriz com todos os caminhos do mapa gerado.

    A partir do centro da matriz, sorteamos varias direções para a qual iremos gerar um caminho no mapa.
    a variável path dita a distância que será percorrida até que o algoritmo escolha outra direção
    para se movimentar. Caso a direção escolhida esteja fora do intervalo da matriz, escolhemos novamente
    a direção até gerar uma posição válida. Então vamos incrementando as posições para movimentar pela
    matriz e as células pelas quais passamos adicionamos o valor 1 para indicar que há um caminho ali.

    Args:
        rows (int): indica a quantidade de linhas da matriz;
        columns (int): indica a quantidade de colunas da matriz.

    Returns:
        room (list[list[int]]): matriz de posições do mapa.
    '''
    
    generator = Xorshift64(seed)

    x_position = columns // 2
    y_position = rows // 2
    
    room =  np.zeros((rows, columns), dtype=int)

    print(seed, steps)

    for _ in range(steps):
        path = 5
        
        x, y = choose_direction(generator)
        
        while (path > 0):
            
            if ((2 <= x_position + x < rows - 2) and (2 <= y_position + y < columns - 2)):
                room[x_position][y_position] = 1

                x_position += x
                y_position += y
            
                path -= 1
            else:
                x, y = choose_direction(generator)

    return room


def room_screen():
    
    '''
    Exibi uma tela com o mapa gerado.

    Criamos um canvas (área de desenho) e configuramos um grid de 60 linhas e 80 colunas
    no qual cada célula será preenchida com uma cor que irá representar um caminho ou não no nosso mapa.

    Args:
        None

    Returns:
        None
    '''

    def draw_room(room):
            
        '''
        Desenha o mapa baseada na matriz gerada pela função generate_room();

        A função pecorre toda a matriz e verifica o valor presente na célular:
        Caso seja 0 preenche o grid com a cor preta, caso seja 1 (caminho) pinta o grid com a cor magenta.

        Args:
            room (list[list[int]]): matriz de posições do mapa.

        Returns:
            None
        '''
 
        for row in range(rows):
            for column in range(columns):

                if (room[row][column]) == 1:
                    canvas.create_rectangle(column * cell, row * cell, (column + 1) * cell, (row + 1) * cell, fill="#FA1E81", outline="#FA1E81")
                else:
                    canvas.create_rectangle(column * cell, row * cell, (column + 1) * cell, (row + 1) * cell, fill="black")


    width = 800
    height = 600
    cell = 10

    rows = height // cell
    columns = width // cell

    window = tk.Tk()
    window.geometry(f'{width}x{height}')
    window.title('Room')

    room = generate_room(rows, columns)
    
    canvas = tk.Canvas(window, width=width, height=height)
    canvas.pack()

    draw_room(room)

    window.mainloop()


def main_screen():

    
    '''
    Tela principal do sistema.

    Há dois campos: um para que o usuário entre com uma seed e outro para que ele entre com o tamanho do mapa.
    O Botão para gerar mapa possui uma função lambda responsável por chamar as funções que pegam os valores dos campos
    e a room_screen().

    Args:
        None

    Returns:
        None

    '''
 
    def get_seed_entry():
    
        '''
        Pega o valor presente no campo seed, caso esteja vazio o algoritmo preenche com o tempo em segundos do computador.

        Args:
            None

        Returns:
            None

        '''
 
        global seed
        
        if (seed_entry.get()):
            seed = seed_entry.get()
        
        else:
            seed = time()

        seed = int(seed)
        
    
    def get_steps_entry():
            
        '''
        Pega o valor do campo steps, caso esteja vazio o algortimo preenche com o valor 20.

        Args:
            None

        Returns:
            None
        '''
 
        global steps

        if (steps_entry.get()):
            steps = steps_entry.get()
        
        else:
            steps = 20

        steps = int(steps)


    width = 400
    height = 300

    window = tk.Tk()
    window.geometry(f'{width}x{height}')
    window.title('Generate Random Rooms')

    label1 = tk.Label(window, text='Choose Seed')
    label1.pack()

    seed_entry = tk.Entry(window)
    seed_entry.pack(pady=10)
    
    label2 = tk.Label(window, text='Choose Steps')
    label2.pack()

    steps_entry = tk.Entry(window)
    steps_entry.pack(pady=10)

    button = tk.Button(window, text='Generate Room', command=lambda: (get_seed_entry(), get_steps_entry(), room_screen()))
    button.pack(pady=10)
    
    read_seed_button = tk.Button(window, text='Display Saved Seeds', command=saved_seeds_screen)
    read_seed_button.pack(pady=10)
    
    save_button = tk.Button(window, text='Save', command=save_screen)
    save_button.pack(pady=10)

    window.mainloop()


def main():
        
    '''
    Função principal.

    Args:
        None

    Returns:
        None
    '''
 
    create_archive()
    main_screen()


if __name__ == '__main__':
    main()
